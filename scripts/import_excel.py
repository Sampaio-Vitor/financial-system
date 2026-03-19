"""Import data from carteira_investimentos Excel spreadsheet into the database."""
import asyncio
import sys
import os
from datetime import datetime
from decimal import Decimal

import openpyxl
from sqlalchemy import select

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend"))

from app.database import AsyncSessionLocal, engine, Base
from app.models.user import User
from app.models.asset import Asset, AssetType
from app.models.purchase import Purchase
from app.models.fixed_income import FixedIncomePosition
from app.models.allocation_target import AllocationTarget
from app.models.settings import UserSettings
from app.models.financial_reserve import FinancialReserveEntry

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "carteira_investimentos (8).xlsx")

TYPE_MAP = {"Stock": AssetType.STOCK, "Acao": AssetType.ACAO, "FII": AssetType.FII}
CLASS_MAP = {
    "Stocks (EUA)": AssetType.STOCK,
    "Acoes (Brasil)": AssetType.ACAO,
    "FIIs": AssetType.FII,
    "Renda Fixa": AssetType.RF,
}


async def import_all():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    async with AsyncSessionLocal() as db:
        # Get user (must exist via seed_user.py first)
        result = await db.execute(select(User).limit(1))
        user = result.scalar_one_or_none()
        if not user:
            print("ERROR: No user found. Run seed_user.py first.")
            return

        # 1. Import assets from 'Ativos' sheet
        assets_imported = await _import_assets(db, wb["Ativos"])
        print(f"Imported {assets_imported} assets")

        # 2. Import purchases from 'Aportes' sheet
        purchases_imported = await _import_purchases(db, wb["Aportes"], user.id)
        print(f"Imported {purchases_imported} purchases")

        # 3. Import RF positions from 'Renda Fixa' sheet
        rf_imported = await _import_fixed_income(db, wb["Renda Fixa"], user.id)
        print(f"Imported {rf_imported} fixed income positions")

        # 4. Import allocation targets from 'Alocacao' sheet
        targets_imported = await _import_allocation(db, wb["Alocacao"], user.id)
        print(f"Imported {targets_imported} allocation targets")

        # 5. Import USD/BRL rate from 'Posicao' sheet
        await _import_settings(db, wb["Posicao"], user.id)
        print("Imported USD/BRL rate")

        # 6. Import financial reserve from 'Posicao' sheet
        reserve_imported = await _import_financial_reserve(db, wb["Posicao"], user.id)
        if reserve_imported:
            print("Imported financial reserve entry")

        await db.commit()
        print("\nImport complete!")


async def _import_assets(db, ws) -> int:
    count = 0
    sections = [
        (6, 35, AssetType.STOCK),   # Stocks rows 6-35
        (39, 68, AssetType.ACAO),   # Acoes rows 39-68
        (72, 81, AssetType.FII),    # FIIs rows 72-81
        (85, 87, AssetType.RF),     # RF rows 85-87
    ]

    for start_row, end_row, asset_type in sections:
        for row in range(start_row, end_row + 1):
            ticker = ws.cell(row=row, column=3).value  # Column C
            description = ws.cell(row=row, column=4).value or ""  # Column D

            if not ticker:
                continue

            ticker = str(ticker).strip().upper()

            # Check if already exists
            existing = await db.execute(select(Asset).where(Asset.ticker == ticker))
            if existing.scalar_one_or_none():
                continue

            asset = Asset(ticker=ticker, type=asset_type, description=str(description).strip())
            db.add(asset)
            count += 1

    await db.flush()
    return count


async def _import_purchases(db, ws, user_id: int) -> int:
    count = 0

    # Build ticker -> asset_id map
    result = await db.execute(select(Asset))
    ticker_to_id = {a.ticker: a.id for a in result.scalars().all()}

    for row in range(5, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        ticker = ws.cell(row=row, column=2).value
        qty = ws.cell(row=row, column=4).value
        unit_price = ws.cell(row=row, column=5).value
        total_value = ws.cell(row=row, column=6).value

        if not date_val or not ticker or not qty:
            continue

        ticker = str(ticker).strip().upper()
        asset_id = ticker_to_id.get(ticker)
        if not asset_id:
            print(f"  Warning: Unknown ticker '{ticker}' in Aportes row {row}, skipping")
            continue

        # Parse date
        if isinstance(date_val, datetime):
            purchase_date = date_val.date()
        else:
            continue

        qty_dec = Decimal(str(qty))
        price_dec = Decimal(str(unit_price)) if unit_price else Decimal("0")
        total_dec = Decimal(str(total_value)) if total_value else qty_dec * price_dec

        # Check for duplicate
        existing = await db.execute(
            select(Purchase).where(
                Purchase.asset_id == asset_id,
                Purchase.purchase_date == purchase_date,
                Purchase.quantity == qty_dec,
                Purchase.user_id == user_id,
            )
        )
        if existing.scalar_one_or_none():
            continue

        purchase = Purchase(
            asset_id=asset_id,
            user_id=user_id,
            purchase_date=purchase_date,
            quantity=qty_dec,
            unit_price=price_dec,
            total_value=total_dec,
        )
        db.add(purchase)
        count += 1

    await db.flush()
    return count


async def _import_fixed_income(db, ws, user_id: int) -> int:
    count = 0

    # Build ticker -> asset_id map for RF
    result = await db.execute(select(Asset).where(Asset.type == AssetType.RF))
    ticker_to_id = {a.ticker: a.id for a in result.scalars().all()}

    for row in range(6, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        rf_type = ws.cell(row=row, column=2).value
        description = ws.cell(row=row, column=3).value
        applied = ws.cell(row=row, column=4).value
        balance = ws.cell(row=row, column=5).value

        if not date_val or not rf_type or not applied:
            continue

        rf_type = str(rf_type).strip()
        asset_id = ticker_to_id.get(rf_type)
        if not asset_id:
            # Try uppercase
            asset_id = ticker_to_id.get(rf_type.upper())
        if not asset_id:
            print(f"  Warning: Unknown RF type '{rf_type}' in row {row}, skipping")
            continue

        if isinstance(date_val, datetime):
            start_date = date_val.date()
        else:
            continue

        yield_val = ws.cell(row=row, column=6).value or 0
        yield_pct = ws.cell(row=row, column=7).value or 0
        maturity = ws.cell(row=row, column=8).value
        maturity_date = maturity.date() if isinstance(maturity, datetime) else None

        fi = FixedIncomePosition(
            asset_id=asset_id,
            user_id=user_id,
            description=str(description or rf_type).strip(),
            start_date=start_date,
            applied_value=Decimal(str(applied)),
            current_balance=Decimal(str(balance)) if balance else Decimal(str(applied)),
            yield_value=Decimal(str(yield_val)),
            yield_pct=Decimal(str(yield_pct)),
            maturity_date=maturity_date,
        )
        db.add(fi)
        count += 1

    await db.flush()
    return count


async def _import_allocation(db, ws, user_id: int) -> int:
    count = 0
    for row in range(7, 11):  # Rows 7-10
        class_name = ws.cell(row=row, column=2).value
        pct = ws.cell(row=row, column=3).value

        if not class_name or not pct:
            continue

        asset_class = CLASS_MAP.get(str(class_name).strip())
        if not asset_class:
            continue

        # Delete existing
        existing = await db.execute(
            select(AllocationTarget).where(
                AllocationTarget.user_id == user_id,
                AllocationTarget.asset_class == asset_class,
            )
        )
        for e in existing.scalars().all():
            await db.delete(e)

        target = AllocationTarget(
            user_id=user_id,
            asset_class=asset_class,
            target_pct=Decimal(str(pct)),
        )
        db.add(target)
        count += 1

    await db.flush()
    return count


async def _import_settings(db, ws, user_id: int):
    rate = ws.cell(row=2, column=3).value  # Cell C2
    if not rate:
        return

    result = await db.execute(select(UserSettings).where(UserSettings.user_id == user_id))
    settings = result.scalar_one_or_none()
    if settings:
        settings.usd_brl_rate = Decimal(str(rate))
        settings.rate_updated_at = datetime.utcnow()
    else:
        settings = UserSettings(
            user_id=user_id,
            usd_brl_rate=Decimal(str(rate)),
            rate_updated_at=datetime.utcnow(),
        )
        db.add(settings)


async def _import_financial_reserve(db, ws, user_id: int) -> bool:
    # Row 89, column E (5) in the Posicao sheet contains the reserve value
    value = ws.cell(row=89, column=5).value
    if not value:
        print("  Warning: No financial reserve value found at Posicao!E89")
        return False

    # Check if a reserve entry already exists for this user
    existing = await db.execute(
        select(FinancialReserveEntry).where(FinancialReserveEntry.user_id == user_id).limit(1)
    )
    if existing.scalar_one_or_none():
        return False

    entry = FinancialReserveEntry(
        user_id=user_id,
        amount=Decimal(str(value)),
        note="Importado da planilha",
        recorded_at=datetime.utcnow(),
    )
    db.add(entry)
    await db.flush()
    return True


if __name__ == "__main__":
    asyncio.run(import_all())
